#!/usr/bin/env python3
"""
new_powerbi_dashboard.py — Complete Power BI Dashboard Automation (CSV/Excel → Finished PBIP)

Guides the user through every phase of creating a Power BI dashboard from raw data:
  1. Preflight checks (deps, project structure, PBID closed)
  2. Interactive data analysis + user preference collection
  3. DAX measure generation and TMDL patching
  4. Visual layout creation via pbir CLI
  5. Theme application with 5 built-in themes
  6. Final verification

Usage:
    python new_powerbi_dashboard.py <project.pbip>               # Existing project
    python new_powerbi_dashboard.py --create <name> <data-file>  # Brand new project
    python new_powerbi_dashboard.py --non-interactive <project.pbip>  # Skip prompts, use defaults

Requires:
    - pbir-cli (installed via pip if missing)
    - Power BI Desktop (PBID) project saved as .pbip
    - Python 3.8+
"""

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import tempfile
import time
from pathlib import Path

# ──────────────────────────────────────────────────────────────
# CONSTANTS
# ──────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_BASE = Path.home() / "AppData/Local/hermes/skills/powerbi/powerbi-orchestrator"
SCRIPTS_DIR = SKILL_BASE / "scripts"
REF_DIR = SKILL_BASE / "references"

VALIDATE_SCRIPT = SCRIPTS_DIR / "validate_pbip.py"
FIX_TMDL_SCRIPT = SCRIPTS_DIR / "fix_tmdl.py"
APPLY_THEME_SCRIPT = SCRIPTS_DIR / "apply_theme.py"
CHECK_OVERLAPS_SCRIPT = SCRIPTS_DIR / "check_overlaps.py"

BUILTIN_THEMES = {
    "1": {
        "id": "slate-terracotta",
        "name": "Slate & Terracotta (Dark — recommended for financial/business)",
        "mode": "dark",
        "colors": {
            "dataColors": ["#A56F63", "#D99B7F", "#3B82F6", "#10B981", "#F59E0B", "#8B5CF6", "#EC4899", "#14B8A6"],
            "background": "#0F3040",
            "foreground": "#F8FAFC",
            "tableAccent": "#A56F63",
            "good": "#10B981",
            "bad": "#EF4444",
            "neutral": "#A56F63",
        },
    },
    "2": {
        "id": "magenta-blossom",
        "name": "Magenta Blossom (Light — recommended for marketing/social)",
        "mode": "light",
        "colors": {
            "dataColors": ["#92003A", "#F62477", "#FFADEE", "#FFE185", "#3B82F6", "#10B981", "#8B5CF6", "#F59E0B"],
            "background": "#FFFFFF",
            "foreground": "#111827",
            "tableAccent": "#92003A",
            "good": "#10B981",
            "bad": "#EF4444",
            "neutral": "#92003A",
        },
    },
    "3": {
        "id": "ecotone-spring",
        "name": "Ecotone Spring (Light — recommended for environmental/nature)",
        "mode": "light",
        "colors": {
            "dataColors": ["#769826", "#A1CB35", "#FFDE4E", "#FF9D4D", "#3B82F6", "#8B5CF6", "#EC4899", "#14B8A6"],
            "background": "#F5F2EB",
            "foreground": "#1a1a2e",
            "tableAccent": "#769826",
            "good": "#A1CB35",
            "bad": "#EF4444",
            "neutral": "#769826",
        },
    },
    "4": {
        "id": "roasted-espresso",
        "name": "Roasted Espresso (Dark — balanced, good default)",
        "mode": "dark",
        "colors": {
            "dataColors": ["#60241E", "#95271D", "#B34A44", "#E77B49", "#3B82F6", "#10B981", "#F59E0B", "#8B5CF6"],
            "background": "#1A0F0D",
            "foreground": "#F8FAFC",
            "tableAccent": "#B34A44",
            "good": "#10B981",
            "bad": "#EF4444",
            "neutral": "#B34A44",
        },
    },
    "5": {
        "id": "vintage-nordic",
        "name": "Vintage Nordic (Light — formal/elegant)",
        "mode": "light",
        "colors": {
            "dataColors": ["#0B1849", "#124D1C", "#E4B028", "#EBEDE3", "#3B82F6", "#8B5CF6", "#EC4899", "#14B8A6"],
            "background": "#EBEDE3",
            "foreground": "#0B1849",
            "tableAccent": "#0B1849",
            "good": "#124D1C",
            "bad": "#C0392B",
            "neutral": "#0B1849",
        },
    },
}

# Default visual layout templates per page configuration
VISUAL_TEMPLATE_1PAGE = [
    # (type, name, title, x, y, w, h, data_fields)
    # Row 1: KPIs
    ("card", "vis_kpi_1", "🔑 Total Sales", 20, 20, 195, 105, []),
    ("card", "vis_kpi_2", "🔑 Average Sale", 225, 20, 195, 105, []),
    ("card", "vis_kpi_3", "🔑 Total Quantity", 430, 20, 195, 105, []),
    # Row 2: Donut + Bar
    ("donutChart", "vis_donut_1", "🍩 Sales by Category", 20, 140, 300, 280, []),
    ("barChart", "vis_bar_1", "📊 Top Values", 330, 140, 300, 280, []),
    # Row 3: Slicer + leftover
    ("slicer", "vis_slicer_1", "", 20, 435, 250, 85, []),
    # Row 4: Line chart
    ("lineChart", "vis_line_1", "📈 Trend Over Time", 20, 535, 610, 390, []),
]

VISUAL_TEMPLATE_2PAGE_P1 = [
    ("card", "vis_kpi_1", "🔑 Total Sales", 20, 20, 195, 105, []),
    ("card", "vis_kpi_2", "🔑 Key Metric 2", 225, 20, 195, 105, []),
    ("card", "vis_kpi_3", "🔑 Key Metric 3", 430, 20, 195, 105, []),
    ("slicer", "vis_slicer_1", "", 20, 140, 250, 85, []),
    ("donutChart", "vis_donut_1", "🍩 Distribution", 20, 240, 300, 280, []),
]

VISUAL_TEMPLATE_2PAGE_P2 = [
    ("card", "vis_kpi_4", "🔑 Additional KPI", 20, 20, 195, 105, []),
    ("barChart", "vis_bar_1", "📊 Detailed Chart", 20, 140, 300, 280, []),
    ("lineChart", "vis_line_1", "📈 Trend", 330, 140, 300, 280, []),
    ("slicer", "vis_slicer_2", "", 20, 435, 250, 85, []),
]

VISUAL_TEMPLATE_3PAGE_P1 = [
    ("card", "vis_kpi_1", "🔑 Key Metric 1", 20, 20, 195, 105, []),
    ("card", "vis_kpi_2", "🔑 Key Metric 2", 225, 20, 195, 105, []),
    ("card", "vis_kpi_3", "🔑 Key Metric 3", 430, 20, 195, 105, []),
    ("slicer", "vis_slicer_1", "", 20, 140, 250, 85, []),
]

VISUAL_TEMPLATE_3PAGE_P2 = [
    ("donutChart", "vis_donut_1", "🍩 Distribution", 20, 20, 300, 280, []),
    ("barChart", "vis_bar_1", "📊 Breakdown", 330, 20, 300, 280, []),
]

VISUAL_TEMPLATE_3PAGE_P3 = [
    ("lineChart", "vis_line_1", "📈 Trend Analysis", 20, 20, 610, 390, []),
]


# ──────────────────────────────────────────────────────────────
# UTILITY FUNCTIONS
# ──────────────────────────────────────────────────────────────

def banner():
    """Print a startup banner."""
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║   Power BI Dashboard Automation — new_powerbi_dashboard.py  ║")
    print("║   CSV/Excel → Finished PBIP Project                        ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()


def step(msg):
    """Print a step header."""
    print(f"\n{'─' * 60}")
    print(f"  🔷 {msg}")
    print(f"{'─' * 60}")


def ok(msg):
    print(f"  ✅ {msg}")


def warn(msg):
    print(f"  ⚠️  {msg}")


def error(msg):
    print(f"  ❌ {msg}")


def info(msg):
    print(f"  ℹ️  {msg}")


def run_cmd(cmd, capture_output=True, check=True, timeout=120):
    """Run a shell command and return result."""
    try:
        result = subprocess.run(
            cmd,
            capture_output=capture_output,
            text=True,
            check=check,
            timeout=timeout,
            shell=isinstance(cmd, str),
        )
        return result
    except subprocess.CalledProcessError as e:
        return e
    except FileNotFoundError:
        return type("Result", (), {
            "returncode": 127,
            "stdout": "",
            "stderr": f"Command not found: {cmd}",
        })()


def confirm(prompt, default="yes"):
    """Ask user a yes/no question."""
    valid = {"yes": True, "y": True, "ye": True, "no": False, "n": False}
    if default is None:
        prompt_text = " [y/n] "
    elif default == "yes":
        prompt_text = " [Y/n] "
    else:
        prompt_text = " [y/N] "
    while True:
        print(f"  ? {prompt}{prompt_text}", end="", flush=True)
        choice = input().strip().lower()
        if default is not None and choice == "":
            return valid[default]
        if choice in valid:
            return valid[choice]
        print(f"  Please respond with 'yes' or 'no' (or 'y' or 'n').")


def choose_option(prompt, options):
    """Present numbered options and let user choose."""
    print(f"\n  ? {prompt}")
    for key, value in options.items():
        print(f"    [{key}] {value['name'] if isinstance(value, dict) else value}")
    print(f"  Enter choice: ", end="", flush=True)
    choice = input().strip()
    while choice not in options:
        print(f"  Invalid choice. Options: {', '.join(options.keys())}")
        print(f"  Enter choice: ", end="", flush=True)
        choice = input().strip()
    return choice


def ask(prompt, default=None):
    """Ask user for text input."""
    if default:
        print(f"  ? {prompt} [{default}]: ", end="", flush=True)
    else:
        print(f"  ? {prompt}: ", end="", flush=True)
    answer = input().strip()
    if not answer and default:
        return default
    return answer


# ──────────────────────────────────────────────────────────────
# PHASE 0: COMMAND-LINE ARGUMENTS
# ──────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Power BI Dashboard Automation — CSV/Excel to Finished PBIP",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s project.pbip                                 # Interactive mode on existing project
  %(prog)s --create MyDashboard data/sales.csv          # Create new project with data file
  %(prog)s --non-interactive project.pbip               # Skip prompts, use smart defaults
  %(prog)s --skip-data-analysis project.pbip             # Skip data file analysis
  %(prog)s --verbose project.pbip                       # Detailed output
        """,
    )
    parser.add_argument("project", nargs="?", help="Path to .pbip project file")
    parser.add_argument("--create", nargs=2, metavar=("NAME", "DATA_FILE"),
                        help="Create a new PBIP project with the given name and data file")
    parser.add_argument("--non-interactive", action="store_true",
                        help="Run in non-interactive mode (use smart defaults)")
    parser.add_argument("--skip-data-analysis", action="store_true",
                        help="Skip the CSV/Excel column analysis phase")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Verbose output")
    parser.add_argument("--theme", choices=list(BUILTIN_THEMES.keys()) + [t["id"] for t in BUILTIN_THEMES.values()],
                        help="Skip theme selection, use this theme directly (by index number or id)")
    parser.add_argument("--pages", type=int, choices=[1, 2, 3], default=None,
                        help="Number of pages (1-3) for the dashboard layout")
    return parser.parse_args()


# ──────────────────────────────────────────────────────────────
# PHASE 1: PREFLIGHT CHECKS
# ──────────────────────────────────────────────────────────────

def check_pbir_cli():
    """Check if pbir CLI is installed, install if not."""
    step("PREFLIGHT: Checking pbir CLI")
    result = run_cmd("which pbir || where pbir", check=False, timeout=30)
    if result.returncode == 0:
        version = run_cmd("pbir --version 2>/dev/null || echo version unknown", check=False)
        ok(f"pbir CLI available: {version.stdout.strip()}")
        return True
    else:
        info("pbir CLI not found. Attempting installation...")
        install = run_cmd("pip install pbir-cli", check=False, timeout=120)
        if install.returncode == 0:
            ok("pbir CLI installed successfully")
            return True
        else:
            error("Failed to install pbir-cli. Install manually: pip install pbir-cli")
            return False


def verify_project_structure(project_root, project_name):
    """Verify the PBIP project has the correct structure."""
    step("PREFLIGHT: Verifying project structure")
    missing = []

    checks = [
        (project_root / f"{project_name}.pbip", ".pbip file"),
        (project_root / f"{project_name}.Report" / "definition.pbir", "definition.pbir"),
        (project_root / f"{project_name}.Report" / "definition" / "report.json", "report.json"),
        (project_root / f"{project_name}.Report" / "definition" / "version.json", "version.json"),
        (project_root / f"{project_name}.Report" / "definition" / "pages" / "pages.json", "pages.json"),
        (project_root / f"{project_name}.SemanticModel" / "definition.pbism", "definition.pbism"),
    ]

    for path, label in checks:
        if not path.exists():
            missing.append(label)
            error(f"Missing: {label} ({path})")

    # Check for TMDL files
    tables_dir = project_root / f"{project_name}.SemanticModel" / "definition" / "tables"
    if tables_dir.exists():
        tmdl_files = list(tables_dir.glob("*.tmdl"))
        if tmdl_files:
            ok(f"Found {len(tmdl_files)} table TMDL file(s)")
        else:
            missing.append("No .tmdl files in tables/")
            error("No .tmdl files in SemanticModel/definition/tables/")
    else:
        missing.append("tables directory")
        error(f"SemanticModel tables dir not found: {tables_dir}")

    # Check for page directories
    pages_dir = project_root / f"{project_name}.Report" / "definition" / "pages"
    if pages_dir.exists():
        page_dirs = [d for d in pages_dir.iterdir() if d.is_dir()]
        if page_dirs:
            ok(f"Found {len(page_dirs)} page(s): {', '.join(p.name for p in page_dirs)}")
        else:
            warn("No page directories found yet (new project)")
    else:
        warn("No pages directory yet")

    return len(missing) == 0


def close_pbid():
    """Ensure Power BI Desktop is closed."""
    step("PREFLIGHT: Ensuring Power BI Desktop is closed")
    result = run_cmd("taskkill /IM PBIDesktop.exe /F 2>/dev/null; echo 'done'", check=False)
    if result.returncode == 0 or "(PID:" in result.stdout or "done" in result.stdout:
        ok("Power BI Desktop closed (or was not running)")
    else:
        ok("Power BI Desktop was not running")


def delete_cache(project_root, project_name):
    """Delete cache.abf if it exists."""
    cache_path = project_root / f"{project_name}.SemanticModel" / ".pbi" / "cache.abf"
    if cache_path.exists():
        cache_path.unlink()
        ok(f"Deleted cache: {cache_path}")
    else:
        ok("cache.abf not found (good)")


def run_validation(project_root, project_name):
    """Run validate_pbip.py and parse results."""
    step("PREFLIGHT: Running structural validation")
    pbip_path = project_root / f"{project_name}.pbip"
    if not VALIDATE_SCRIPT.exists():
        error(f"validate_pbip.py not found at {VALIDATE_SCRIPT}")
        # Try a direct check instead
        return basic_validation(project_root, project_name)

    result = run_cmd(f'python "{VALIDATE_SCRIPT}" "{pbip_path}"', check=False)
    if result.returncode == 0:
        ok("Structural validation passed")
        # Parse JSON from output
        try:
            data = json.loads(result.stdout)
            if data.get("warnings", 0) > 0:
                warn(f"Validation warnings: {data['warnings']}")
            return True
        except (json.JSONDecodeError, KeyError):
            warn("Could not parse validation output, but exit code was OK")
            return True
    else:
        error("Structural validation failed")
        try:
            data = json.loads(result.stdout)
            for c in data.get("checks", []):
                if c.get("status") == "error":
                    error(f"  [{c['check']}] {c['message']}")
        except json.JSONDecodeError:
            warn("Could not parse validation output")
            print(result.stderr)
        return False


def basic_validation(project_root, project_name):
    """Simple validation when validate_pbip.py is unavailable."""
    ok_count = 0
    fail_count = 0
    checks = [
        (project_root / f"{project_name}.pbip", ".pbip file"),
        (project_root / f"{project_name}.Report" / "definition.pbir", "definition.pbir"),
        (project_root / f"{project_name}.Report" / "definition" / "report.json", "report.json"),
    ]
    for path, label in checks:
        if path.exists():
            ok_count += 1
            ok(f"Found: {label}")
        else:
            fail_count += 1
            error(f"Missing: {label}")
    return fail_count == 0


# ──────────────────────────────────────────────────────────────
# PHASE 2: DATA FILE ANALYSIS
# ──────────────────────────────────────────────────────────────

def auto_detect_type(value):
    """Detect column type from a sample value."""
    if value is None or value.strip() == "":
        return None
    v = value.strip()
    # Try integer
    try:
        int(v)
        return "integer"
    except ValueError:
        pass
    # Try float
    try:
        float(v.replace(",", ""))
        return "decimal"
    except ValueError:
        pass
    # Try date
        date_patterns = [
            r"^\d{4}-\d{2}-\d{2}$", r"^\d{2}/\d{2}/\d{4}$",
            r"^\d{2}-\d{2}-\d{4}$", r"^\d{4}/\d{2}/\d{2}$",
            r"^\d{2}/\d{2}/\d{2}$",
        ]
    for pat in date_patterns:
        if re.match(pat, v):
            return "date"
    # Try boolean
    if v.lower() in ("true", "false", "yes", "no", "1", "0"):
        return "boolean"
    # Default to text
    return "text"


def analyze_csv(file_path, sample_size=100):
    """Analyze a CSV file and return column info."""
    step(f"ANALYZING CSV: {file_path}")
    info(f"Reading first {sample_size} rows to detect schema...")
    try:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            if not headers:
                error("CSV has no headers")
                return None

            rows = []
            for i, row in enumerate(reader):
                if i >= sample_size:
                    break
                rows.append(row)

            total_rows = 0
            # Count total rows
            with open(file_path, "r", encoding="utf-8-sig") as fc:
                total_rows = sum(1 for _ in fc) - 1  # minus header

            columns = {}
            for h in headers:
                values = [r.get(h, "") for r in rows if r.get(h, "").strip()]
                if not values:
                    columns[h] = {"type": "text", "sample": "", "unique": 0,
                                  "null_pct": 100, "is_numeric": False, "is_date": False}
                    continue
                # Detect type from non-empty values
                types = [auto_detect_type(v) for v in values]
                types = [t for t in types if t is not None]
                if not types:
                    col_type = "text"
                else:
                    # Use most specific type
                    if all(t in ("integer", "decimal") for t in types):
                        # Check if all integers
                        if all(t == "integer" for t in types):
                            col_type = "integer"
                        else:
                            col_type = "decimal"
                    elif "date" in types:
                        col_type = "date"
                    elif "boolean" in types:
                        col_type = "boolean"
                    else:
                        col_type = "text"

                unique_vals = len(set(values))
                null_count = sum(1 for r in rows if not r.get(h, "").strip())
                null_pct = round(null_count / len(rows) * 100 if rows else 0, 1)

                columns[h] = {
                    "type": col_type,
                    "sample": values[0] if values else "",
                    "unique": unique_vals,
                    "null_pct": null_pct,
                    "is_numeric": col_type in ("integer", "decimal"),
                    "is_date": col_type == "date",
                }

            info(f"  Total rows in file: ~{max(total_rows, len(rows))}")
            info(f"  Columns detected: {len(headers)}")
            for h in headers:
                type_icon = {"integer": "#️⃣", "decimal": "🔢", "date": "📅", "text": "📝", "boolean": "✅"}
                icon = type_icon.get(columns[h]["type"], "❓")
                info(f"    {icon} {h}: {columns[h]['type']} "
                     f"(unique={columns[h]['unique']}, null={columns[h]['null_pct']}%)")

            return {"headers": headers, "columns": columns, "rows": rows, "total_rows": total_rows}
    except Exception as e:
        error(f"Failed to analyze CSV: {e}")
        return None


def analyze_excel(file_path, sample_size=100):
    """Analyze an Excel file. We read via csv if we can convert, otherwise advise."""
    step(f"ANALYZING EXCEL: {file_path}")
    warn("Full Excel analysis requires openpyxl or pandas. Attempting basic analysis...")

    ext = Path(file_path).suffix.lower()
    # Try to read the file — Excel files can sometimes be read as CSV with xlrd
    # For simplicity, we'll fall back to asking the user to convert to CSV, or
    # if they have Python tools available, use them
    try:
        import csv  # always available
        # Excel isn't directly readable by csv module; prompt user
        info("For Excel files, we recommend:")
        info("  1. Open the file in Excel or a text editor")
        info("  2. Save As → CSV (Comma delimited) (*.csv)")
        info("  3. Run this script again with the CSV file path")
        info("\n  Alternatively, the script can attempt to use available Python libraries...")

        # Try openpyxl if available
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_col=ws.max_column))]
            headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(headers)]

            rows = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if i >= sample_size:
                    break
                row_dict = {}
                for j, val in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = str(val) if val is not None else ""
                rows.append(row_dict)

            columns = {}
            for h in headers:
                values = [r.get(h, "") for r in rows if r.get(h, "").strip()]
                if not values:
                    columns[h] = {"type": "text", "unique": 0, "null_pct": 100,
                                  "is_numeric": False, "is_date": False}
                    continue
                types = [auto_detect_type(v) for v in values]
                types = [t for t in types if t is not None]
                if all(t in ("integer", "decimal") for t in types):
                    col_type = "decimal" if "decimal" in types else "integer"
                elif "date" in types:
                    col_type = "date"
                elif "boolean" in types:
                    col_type = "boolean"
                else:
                    col_type = "text"
                unique_vals = len(set(values))
                null_count = sum(1 for r in rows if not r.get(h, "").strip())
                null_pct = round(null_count / len(rows) * 100 if rows else 0, 1)
                columns[h] = {"type": col_type, "unique": unique_vals, "null_pct": null_pct,
                              "is_numeric": col_type in ("integer", "decimal"),
                              "is_date": col_type == "date"}

            wb.close()
            info(f"  Columns detected: {len(headers)}")
            for h in headers:
                type_icon = {"integer": "#️⃣", "decimal": "🔢", "date": "📅", "text": "📝", "boolean": "✅"}
                icon = type_icon.get(columns[h]["type"], "❓")
                info(f"    {icon} {h}: {columns[h]['type']} (unique={columns[h]['unique']}, null={columns[h]['null_pct']}%)")

            return {"headers": headers, "columns": columns, "rows": rows, "total_rows": 0}

        except ImportError:
            error("openpyxl not available. Can't parse Excel directly.")
            info("Install with: pip install openpyxl")
            info("Or convert your Excel file to CSV and re-run.")
            return None

    except Exception as e:
        error(f"Failed to analyze Excel: {e}")
        return None


def analyze_data_file(file_path):
    """Analyze a CSV or Excel file."""
    p = Path(file_path)
    if not p.exists():
        error(f"Data file not found: {file_path}")
        return None

    ext = p.suffix.lower()
    if ext == ".csv":
        return analyze_csv(file_path)
    elif ext in (".xlsx", ".xls"):
        return analyze_excel(file_path)
    else:
        error(f"Unsupported file format: {ext}. Use .csv or .xlsx")
        return None


def suggest_dimensions_measures(analysis, purpose="analysis"):
    """Based on column types, suggest which are dimensions and which are measures."""
    if not analysis:
        return [], []

    dimensions = []
    measures = []
    date_cols = []

    for h, info in analysis["columns"].items():
        if info["is_numeric"]:
            measures.append(h)
        elif info["is_date"]:
            date_cols.append(h)
            dimensions.append(h)
        elif info["type"] == "boolean":
            dimensions.append(h)
        else:
            # Text column: if low cardinality -> dimension, otherwise could be label
            if info["unique"] <= 25:
                dimensions.append(h)
            else:
                # High cardinality text — could be a label or an ID
                dimensions.append(h)

    return dimensions, measures, date_cols


# ──────────────────────────────────────────────────────────────
# PHASE 3: INTERACTIVE USER INPUT
# ──────────────────────────────────────────────────────────────

def collect_user_preferences(args, analysis=None):
    """Collect user preferences interactively or from args."""
    prefs = {
        "data_file": None,
        "purpose": "analysis",
        "audience": "executives",
        "theme_choice": None,
        "num_pages": 1,
        "custom_visuals": [],
    }

    if args.non_interactive:
        info("Non-interactive mode: using smart defaults")
        prefs["purpose"] = "analysis"
        prefs["audience"] = "executives"
        prefs["num_pages"] = args.pages if args.pages else 1
        if args.theme:
            prefs["theme_choice"] = resolve_theme(args.theme)
        else:
            prefs["theme_choice"] = "4"  # roasted-espresso as default
        return prefs

    step("COLLECTING USER PREFERENCES")

    if args.create:
        prefs["data_file"] = args.create[1]
        info(f"Data file from --create: {prefs['data_file']}")

    if not prefs["data_file"]:
        prefs["data_file"] = ask("Path to your data file (.csv or .xlsx)")

    prefs["purpose"] = ask("Dashboard purpose (e.g. sales analysis, performance, inventory)",
                           default="analysis")
    prefs["audience"] = ask("Target audience (e.g. executives, operations, external client)",
                            default="executives")

    # Theme selection
    print(f"\n  ? Select a color theme:")
    for k, v in BUILTIN_THEMES.items():
        mode_indicator = "🌙" if v["mode"] == "dark" else "☀️"
        print(f"    [{k}] {mode_indicator} {v['name']}")
    if args.theme:
        choice = resolve_theme(args.theme)
        info(f"Theme specified via --theme: {choice}")
    else:
        # Recommend based on purpose
        purpose_lower = prefs["purpose"].lower()
        if "finance" in purpose_lower or "sale" in purpose_lower or "revenue" in purpose_lower:
            recommend = "1"
            info(f"  💡 Recommendation: Slate & Terracotta (Dark) — great for financial data")
        elif "market" in purpose_lower or "social" in purpose_lower or "campaign" in purpose_lower:
            recommend = "2"
            info(f"  💡 Recommendation: Magenta Blossom (Light) — great for marketing data")
        elif "environment" in purpose_lower or "nature" in purpose_lower or "ecolog" in purpose_lower:
            recommend = "3"
            info(f"  💡 Recommendation: Ecotone Spring (Light) — great for nature/environmental data")
        else:
            recommend = "4"
            info(f"  💡 Recommendation: Roasted Espresso (Dark) — balanced default")

        choice = ask(f"Your choice (or press Enter for default [{recommend}])",
                     default=recommend)
    prefs["theme_choice"] = choice if choice in BUILTIN_THEMES else recommend

    # Page count
    if args.pages:
        prefs["num_pages"] = args.pages
        info(f"Page count from --pages: {prefs['num_pages']}")
    else:
        prefs["num_pages"] = int(ask("Number of pages (1-3)", default="1"))

    # Specific visual types?
    has_specific_visuals = confirm("Do you have specific visual types in mind (beyond the defaults)?", default="no")
    if has_specific_visuals:
        info("Enter additional visual types one per line (card, barChart, lineChart, donutChart, slicer, table, matrix)")
        info("Leave blank when done:")
        while True:
            vt = input("    > ").strip()
            if not vt:
                break
            prefs["custom_visuals"].append(vt)

    return prefs


def resolve_theme(theme_val):
    """Resolve theme argument to a key in BUILTIN_THEMES."""
    if theme_val in BUILTIN_THEMES:
        return theme_val
    for k, v in BUILTIN_THEMES.items():
        if v["id"] == theme_val:
            return k
    return "4"  # default


# ──────────────────────────────────────────────────────────────
# PHASE 4: DAX MEASURE GENERATION
# ──────────────────────────────────────────────────────────────

def suggest_measures_for_columns(analysis, purpose):
    """Suggest common DAX measures based on data analysis and dashboard purpose."""
    if not analysis:
        return []

    measures = []
    numeric_cols = [h for h in analysis["headers"] if analysis["columns"][h]["is_numeric"]]
    text_cols = [h for h in analysis["headers"] if not analysis["columns"][h]["is_numeric"]
                 and not analysis["columns"][h]["is_date"]]
    date_cols = [h for h in analysis["headers"] if analysis["columns"][h]["is_date"]]

    purpose_lower = purpose.lower()

    # Always add Total for numeric columns
    for col in numeric_cols[:3]:
        clean_name = col.replace(" ", " ").replace("%", "Pct").replace(".", " ")
        measure_name = f"Total {clean_name}"
        # Sanitize column name for DAX
        safe_col = col.replace("'", "''")
        measures.append({
            "name": measure_name,
            "expression": f"SUM('Table'{'[{}]'.format(safe_col)})",
            "format_string": "$#,##0",
            "description": f"Sum of {col}",
        })

    # Average for the first numeric column
    if numeric_cols:
        col = numeric_cols[0]
        safe_col = col.replace("'", "''")
        measures.append({
            "name": f"Average {col.replace('%', 'Pct')}",
            "expression": f"AVERAGE('Table'{'[{}]'.format(safe_col)})",
            "format_string": "$#,##0.00",
            "description": f"Average of {col}",
        })

    # Count rows
    measures.append({
        "name": "Total Rows",
        "expression": "COUNTROWS('Table')",
        "format_string": "#,##0",
        "description": "Total number of rows",
    })

    # Purpose-specific measures
    if "sale" in purpose_lower or "revenue" in purpose_lower or "finance" in purpose_lower:
        if numeric_cols:
            col = numeric_cols[0]
            safe_col = col.replace("'", "''")
            measures.append({
                "name": "Max Transaction",
                "expression": f"MAX('Table'{'[{}]'.format(safe_col)})",
                "format_string": "$#,##0",
                "description": f"Maximum {col}",
            })
            measures.append({
                "name": "Min Transaction",
                "expression": f"MIN('Table'{'[{}]'.format(safe_col)})",
                "format_string": "$#,##0",
                "description": f"Minimum {col}",
            })

    if "perform" in purpose_lower:
        measures.append({
            "name": "Completion Pct",
            "expression": "DIVIDE(COUNTROWS('Table'), CALCULATE(COUNTROWS('Table'), ALL('Table'))) * 100",
            "format_string": "0.00\"%\"",
            "description": "Completion percentage",
        })

    if "inventory" in purpose_lower or "stock" in purpose_lower:
        if numeric_cols:
            col = numeric_cols[0]
            safe_col = col.replace("'", "''")
            measures.append({
                "name": "Stock Turnover",
                "expression": f"DIVIDE(SUM('Table'{'[{}]'.format(safe_col)}), COUNTROWS('Table'))",
                "format_string": "0.00",
                "description": "Average stock turnover",
            })

    if len(numeric_cols) >= 2:
        c1 = numeric_cols[0].replace("'", "''")
        c2 = numeric_cols[1].replace("'", "''")
        measures.append({
            "name": "Ratio",
            "expression": f"DIVIDE(SUM('Table'{'[{}]'.format(c1)}), SUM('Table'{'[{}]'.format(c2)}))",
            "format_string": "0.00\"%\"",
            "description": f"Ratio of {numeric_cols[0]} to {numeric_cols[1]}",
        })

    return measures


def read_table_tmdl(project_root, project_name):
    """Read TMDL files to find table names and columns."""
    tables_dir = project_root / f"{project_name}.SemanticModel" / "definition" / "tables"
    if not tables_dir.exists():
        error(f"Tables directory not found: {tables_dir}")
        return None, None

    tmdl_files = sorted(tables_dir.glob("*.tmdl"))
    if not tmdl_files:
        error("No TMDL files found")
        return None, None

    table_name = None
    columns = []

    for tf in tmdl_files:
        content = tf.read_text(encoding="utf-8")
        # Extract table name (first line after model/table keyword)
        for line in content.splitlines():
            line = line.strip()
            if line.startswith("table ") or line.startswith("table  "):
                # Extract quoted name: table 'Sales' {
                m = re.match(r"table\s+'([^']+)'", line)
                if m:
                    table_name = m.group(1)
                elif "'" not in line:
                    m2 = re.match(r"table\s+(\w+)", line)
                    if m2:
                        table_name = m2.group(1)
            # Extract column names
            if "column '" in line:
                m = re.match(r".*'([^']+)'.*", line)
                if m:
                    col_name = m.group(1)
                    columns.append(col_name)

        if table_name:
            info(f"Found table '{table_name}' with {len(columns)} column(s) in {tf.name}")
            return table_name, columns

    if not table_name:
        warn("Could not determine table name from TMDL; using 'Table' as fallback")
        return "Table", columns

    return table_name, columns


def add_measure_to_tmdl_direct(tmdl_path, measure_name, expression, format_string=None):
    """Directly patch a TMDL file to add a DAX measure."""
    info(f"Adding measure '{measure_name}' to {tmdl_path.name}...")

    tmdl_path = Path(tmdl_path)
    if not tmdl_path.exists():
        error(f"TMDL file not found: {tmdl_path}")
        return False

    content = tmdl_path.read_text(encoding="utf-8")

    # Sanitize name: replace % with Pct
    safe_name = measure_name.replace("%", "Pct")

    # Build the measure block
    fs_line = f'            formatString: "{format_string}"' if format_string else ""
    data_category = ""
    measure_block = f"""
    measure '{safe_name}':
        {expression}
        {fs_line}
"""

    # Insert before the last closing brace of the table
    # Find the last occurrence of '}' that closes the table
    lines = content.split("\n")
    # Walk backwards to find the table-closing '}'
    insert_pos = None
    for i in range(len(lines) - 1, -1, -1):
        stripped = lines[i].strip()
        if stripped == "}":
            insert_pos = i
            break

    if insert_pos is None:
        error("Could not find insertion point in TMDL (no closing brace)")
        return False

    # Insert the measure block
    lines.insert(insert_pos, measure_block)
    new_content = "\n".join(lines)

    # Ensure LF endings
    new_content = new_content.replace("\r\n", "\n")

    tmdl_path.write_text(new_content, encoding="utf-8")
    ok(f"Measure '{safe_name}' added to {tmdl_path.name}")
    return True


def generate_and_add_measures(project_root, project_name, analysis, purpose, non_interactive):
    """Suggest measures, ask user, and add them to TMDL."""
    step("PHASE 4: DAX MEASURE GENERATION")

    # Find table name and TMDL file
    table_name, existing_cols = read_table_tmdl(project_root, project_name)
    if not table_name:
        error("Cannot proceed with measures without a valid table. Open PBID to load data first.")
        return False

    # Find the TMDL file path
    tables_dir = project_root / f"{project_name}.SemanticModel" / "definition" / "tables"
    tmdl_files = sorted(tables_dir.glob("*.tmdl"))
    if not tmdl_files:
        error("No TMDL files to patch")
        return False

    # Suggest measures
    suggested = suggest_measures_for_columns(analysis, purpose)

    if not suggested:
        info("No measures suggested. Using basic defaults.")
        suggested = [
            {"name": "Total Rows", "expression": "COUNTROWS('{}')".format(table_name),
             "format_string": "#,##0", "description": "Total row count"},
        ]

    # Display suggestions
    info(f"Suggested DAX measures for table '{table_name}':")
    for i, m in enumerate(suggested, 1):
        info(f"  {i}. {m['name']} = {m['expression']}  [{m['description']}]")

    if not non_interactive:
        if not confirm("Add all suggested measures?", default="yes"):
            info("Skipping measure generation. You can add measures manually later.")
            return True
    else:
        info("Non-interactive: adding all suggested measures")

    # Add each measure to the first TMDL file (table)
    # We add to the file that matches the detected table, or the first one
    tmdl_target = tmdl_files[0]

    if table_name and len(tmdl_files) > 1:
        # Try to find matching table
        for tf in tmdl_files:
            content = tf.read_text(encoding="utf-8")
            if f"table '{table_name}'" in content or f"table {table_name}" in content:
                tmdl_target = tf
                break

    for m in suggested:
        # Adjust expression to use actual table name
        expr = m["expression"].replace("'Table'", f"'{table_name}'")
        success = add_measure_to_tmdl_direct(tmdl_target, m["name"], expr, m.get("format_string"))
        if not success:
            warn(f"Failed to add measure '{m['name']}'")

    # Run fix_tmdl.py
    run_fix_tmdl(project_root, project_name)

    return True


# ──────────────────────────────────────────────────────────────
# PHASE 5: VISUAL LAYOUT
# ──────────────────────────────────────────────────────────────

def get_page_guid(page_path):
    """Extract page GUID from page.json."""
    if not page_path.exists():
        return None
    try:
        with open(page_path, "r", encoding="utf-8") as f:
            page = json.load(f)
        return page.get("name", page_path.parent.name)
    except (json.JSONDecodeError, KeyError):
        return page_path.parent.name


def get_page_info(project_root, project_name):
    """Get page GUIDs from the project."""
    pages_dir = project_root / f"{project_name}.Report" / "definition" / "pages"
    if not pages_dir.exists():
        return []

    pages = []
    pages_json = pages_dir / "pages.json"
    if pages_json.exists():
        try:
            with open(pages_json, "r", encoding="utf-8") as f:
                pg = json.load(f)
            for p in pg.get("pageOrder", []):
                pages.append(p)
        except (json.JSONDecodeError, KeyError):
            pass

    # Also scan directories
    for d in sorted(pages_dir.iterdir()):
        if d.is_dir():
            guid = d.name
            if guid not in pages:
                pages.append(guid)

    return pages


def name_page(project_root, project_name, page_guid, new_name):
    """Rename a page by modifying page.json."""
    page_json = project_root / f"{project_name}.Report" / "definition" / "pages" / page_guid / "page.json"
    if not page_json.exists():
        warn(f"page.json not found for {page_guid}")
        return
    try:
        with open(page_json, "r", encoding="utf-8") as f:
            page = json.load(f)
        page["displayName"] = new_name
        with open(page_json, "w", encoding="utf-8") as f:
            json.dump(page, f, indent=2, ensure_ascii=False)
        ok(f"Page renamed to '{new_name}'")
    except Exception as e:
        error(f"Failed to rename page: {e}")


def create_visuals(project_root, project_name, num_pages, analysis, purpose, non_interactive):
    """Create visuals using pbir CLI."""
    step("PHASE 5: VISUAL LAYOUT")

    pages = get_page_info(project_root, project_name)
    if not pages:
        error("No pages found. Have you saved the project in PBID first?")
        return False

    info(f"Found {len(pages)} page(s): {', '.join(p[:8] + '...' if len(p) > 8 else p for p in pages)}")

    # Name pages and assign visual templates
    if num_pages == 1:
        page_names = ["Dashboard"]
        templates = [VISUAL_TEMPLATE_1PAGE]
    elif num_pages == 2:
        page_names = ["Overview", "Details"]
        templates = [VISUAL_TEMPLATE_2PAGE_P1, VISUAL_TEMPLATE_2PAGE_P2]
    else:
        page_names = ["Overview", "Breakdown", "Trends"]
        templates = [VISUAL_TEMPLATE_3PAGE_P1, VISUAL_TEMPLATE_3PAGE_P2, VISUAL_TEMPLATE_3PAGE_P3]

    # Determine what analysis dimension to use
    dimension_col = None
    measure_col = None
    date_col = None

    if analysis:
        dims, measures, dates = suggest_dimensions_measures(analysis, purpose)
        dimension_col = dims[0] if dims else None
        measure_col = measures[0] if measures else None
        date_col = dates[0] if dates else None

    for page_idx, (page_guid, page_name) in enumerate(zip(pages[:num_pages], page_names)):
        info(f"\n  Setting up page: '{page_name}'")
        name_page(project_root, project_name, page_guid, page_name)

        page_path = f"{project_name}.Report/{page_guid}.Page"
        template = templates[page_idx] if page_idx < len(templates) else VISUAL_TEMPLATE_1PAGE

        for vt in template:
            vtype, vname, vtitle, vx, vy, vw, vh = vt[:7]

            # Build data field args
            data_args = []
            if vtype == "card":
                if measure_col:
                    data_args.extend(["-d", f"Values:{measure_col}"])
            elif vtype in ("donutChart", "barChart"):
                if dimension_col:
                    data_args.extend(["-d", f"Category:{dimension_col}"])
                if measure_col:
                    data_args.extend(["-d", f"Y:{measure_col}"])
            elif vtype == "lineChart":
                if date_col:
                    data_args.extend(["-d", f"Category:{date_col}"])
                elif dimension_col:
                    data_args.extend(["-d", f"Category:{dimension_col}"])
                if measure_col:
                    data_args.extend(["-d", f"Y:{measure_col}"])
            elif vtype == "slicer":
                if dimension_col:
                    data_args.extend(["-d", f"Values:{dimension_col}"])

            if not data_args and vtype != "slicer":
                warn(f"  Skipping {vtype} '{vname}' — no suitable data fields")
                continue

            title_arg = ["-t", vtitle] if vtitle else []

            cmd = [
                "pbir", "add", "visual", vtype,
                page_path,
                "-n", vname,
                *title_arg,
                "-x", str(vx),
                "-y", str(vy),
                "-w", str(vw),
                "-h", str(vh),
                *data_args,
            ]

            info(f"  Creating {vtype}: {vname}")
            result = run_cmd(cmd, check=False, timeout=30)
            if result.returncode == 0:
                ok(f"  Created {vtype}: {vname}")
            else:
                err_msg = result.stderr.strip() or result.stdout.strip() or "unknown error"
                warn(f"  Failed to create {vtype} '{vname}': {err_msg}")

    # Run overlap check
    run_check_overlaps(project_root, project_name)

    return True


# ──────────────────────────────────────────────────────────────
# PHASE 6: THEME APPLICATION
# ──────────────────────────────────────────────────────────────

def apply_theme(project_root, project_name, theme_choice):
    """Apply the chosen theme using apply_theme.py or direct file manipulation."""
    step("PHASE 6: THEME & STYLING")

    theme = BUILTIN_THEMES.get(theme_choice)
    if not theme:
        error(f"Invalid theme choice: {theme_choice}")
        return False

    theme_id = theme["id"]
    theme_colors = theme["colors"]

    info(f"Applying theme: {theme['name']}")

    # Try apply_theme.py first
    pbip_path = project_root / f"{project_name}.pbip"
    if APPLY_THEME_SCRIPT.exists():
        info("Using apply_theme.py...")
        result = run_cmd(
            f'python "{APPLY_THEME_SCRIPT}" "{pbip_path}" --theme {theme_id}',
            check=False, timeout=30
        )
        if result.returncode == 0:
            ok(f"Theme '{theme_id}' applied successfully via apply_theme.py")
            return True
        else:
            warn(f"apply_theme.py returned error: {result.stderr.strip()[:200]}")
            warn("Falling back to direct theme application...")
    else:
        warn(f"apply_theme.py not found at {APPLY_THEME_SCRIPT}, applying directly...")

    # Direct application fallback
    # 1. Overwrite theme file
    theme_dirs = [
        project_root / f"{project_name}.Report" / "StaticResources" / "SharedResources" / "BaseThemes",
        project_root / f"{project_name}.Report" / "definition" / "BaseThemes",
    ]

    theme_applied = False
    for td in theme_dirs:
        if td.exists():
            existing = list(td.glob("*.json"))
            if existing:
                target = existing[0]
                theme_json = {
                    "name": "CY26SU05",
                    "dataColors": theme_colors["dataColors"],
                    "background": theme_colors["background"],
                    "foreground": theme_colors["foreground"],
                    "tableAccent": theme_colors["tableAccent"],
                    "good": theme_colors["good"],
                    "bad": theme_colors["bad"],
                    "neutral": theme_colors["neutral"],
                }
                # Preserve original name
                try:
                    with open(target, "r", encoding="utf-8") as f:
                        old = json.load(f)
                    theme_json["name"] = old.get("name", "CY26SU05")
                except Exception:
                    pass

                with open(target, "w", encoding="utf-8") as f:
                    json.dump(theme_json, f, indent=2, ensure_ascii=False)
                ok(f"Theme file overwritten: {target}")
                theme_applied = True

    if not theme_applied:
        warn("No existing theme file found to overwrite")

    # 2. Apply page backgrounds
    pages_dir = project_root / f"{project_name}.Report" / "definition" / "pages"
    if pages_dir.exists():
        bg_color = theme_colors["background"]
        count = 0
        for page_dir in sorted(pages_dir.iterdir()):
            if not page_dir.is_dir():
                continue
            page_json = page_dir / "page.json"
            if not page_json.exists():
                continue
            try:
                with open(page_json, "r", encoding="utf-8") as f:
                    page = json.load(f)
                page.setdefault("objects", {})["background"] = [{
                    "properties": {
                        "color": {"solid": {"color": {"expr": {"Literal": {"Value": f"'{bg_color}'"}}}}},
                        "transparency": {"expr": {"Literal": {"Value": "0"}}},
                    }
                }]
                with open(page_json, "w", encoding="utf-8") as f:
                    json.dump(page, f, indent=2, ensure_ascii=False)
                count += 1
            except Exception as e:
                warn(f"Failed to set background for {page_dir.name}: {e}")
        ok(f"Background applied to {count} page(s)")

    # 3. Apply dataPoint colors to visuals
    visual_files = []
    for vpath in pages_dir.rglob("visuals/*/visual.json"):
        visual_files.append(vpath)

    if visual_files:
        dc = theme_colors["dataColors"]
        updated = 0
        for vpath in visual_files:
            try:
                with open(vpath, "r", encoding="utf-8") as f:
                    v = json.load(f)
                vtype = v.get("visual", {}).get("visualType", "")
                if vtype in ("treemap", "card", "slicer"):
                    continue
                data_point = []
                for i, color in enumerate(dc):
                    data_point.append({
                        "properties": {
                            "fill": {
                                "solid": {
                                    "color": {
                                        "expr": {"ThemeDataColor": {"ColorId": i, "Percent": 0}}
                                    }
                                }
                            }
                        }
                    })
                v.setdefault("visual", {}).setdefault("objects", {})["dataPoint"] = data_point
                with open(vpath, "w", encoding="utf-8") as f:
                    json.dump(v, f, indent=2, ensure_ascii=False)
                updated += 1
            except Exception as e:
                warn(f"Failed to apply colors to {vpath}: {e}")
        ok(f"Data point colors applied to {updated} visual(s)")

    return True


# ──────────────────────────────────────────────────────────────
# VERIFICATION HELPERS
# ──────────────────────────────────────────────────────────────

def run_fix_tmdl(project_root, project_name):
    """Run fix_tmdl.py on the semantic model."""
    step("VERIFICATION: Running fix_tmdl.py")
    sm_def = project_root / f"{project_name}.SemanticModel" / "definition"
    if FIX_TMDL_SCRIPT.exists() and sm_def.exists():
        result = run_cmd(f'python "{FIX_TMDL_SCRIPT}" "{sm_def}"', check=False)
        if result.returncode == 0:
            try:
                data = json.loads(result.stdout)
                if data.get("files_fixed", 0) > 0:
                    ok(f"TMDL fixes applied to {data['files_fixed']} file(s)")
                else:
                    ok("All TMDL files OK")
            except json.JSONDecodeError:
                ok("fix_tmdl.py ran successfully")
        else:
            warn("fix_tmdl.py reported issues (exit code != 0)")
        return True
    else:
        warn("Cannot run fix_tmdl.py (script or path not found)")
        return False


def run_check_overlaps(project_root, project_name):
    """Run check_overlaps.py."""
    pbip_path = project_root / f"{project_name}.pbip"
    if CHECK_OVERLAPS_SCRIPT.exists():
        result = run_cmd(f'python "{CHECK_OVERLAPS_SCRIPT}" "{pbip_path}"', check=False)
        if result.returncode == 0:
            ok("No overlapping visuals detected")
        else:
            try:
                data = json.loads(result.stdout)
                for issue in data.get("issues", []):
                    if issue.get("severity") == "error":
                        error(f"  Overlap: {issue['message']}")
                    else:
                        warn(f"  {issue['message']}")
            except json.JSONDecodeError:
                warn("Overlap check reported issues")
        return True
    else:
        warn("check_overlaps.py not available")
        return False


def final_verification(project_root, project_name):
    """Run all verification checks."""
    step("PHASE 7: FINAL VERIFICATION")

    all_ok = True

    # 1. Structural validation
    if not run_validation(project_root, project_name):
        all_ok = False

    # 2. TMDL fix
    run_fix_tmdl(project_root, project_name)

    # 3. Overlap check
    run_check_overlaps(project_root, project_name)

    # 4. Check CRLF in TMDL
    tables_dir = project_root / f"{project_name}.SemanticModel" / "definition" / "tables"
    if tables_dir.exists():
        crlf_found = False
        for tf in tables_dir.glob("*.tmdl"):
            content = tf.read_bytes()
            if b"\r\n" in content:
                warn(f"CRLF detected in {tf.name}")
                crlf_found = True
                all_ok = False
        if not crlf_found:
            ok("No CRLF in TMDL files")

    # 5. Verify cache.abf deleted
    cache_path = project_root / f"{project_name}.SemanticModel" / ".pbi" / "cache.abf"
    if cache_path.exists():
        warn("cache.abf still exists — will delete now")
        try:
            cache_path.unlink()
            ok("cache.abf deleted")
        except Exception as e:
            error(f"Could not delete cache.abf: {e}")
            all_ok = False
    else:
        ok("cache.abf not present")

    # 6. Verify UTF-8 without BOM on JSON files
    bom_found = False
    for json_file in project_root.rglob("*.json"):
        if json_file.stat().st_size > 0:
            header = json_file.read_bytes()[:3]
            if header == b"\xef\xbb\xbf":
                warn(f"BOM detected in {json_file.relative_to(project_root)}")
                bom_found = True
                all_ok = False
    if not bom_found:
        ok("No UTF-8 BOM in JSON files")

    return all_ok


# ──────────────────────────────────────────────────────────────
# FINAL OUTPUT
# ──────────────────────────────────────────────────────────────

def print_final_instructions(project_root, project_name):
    """Print final instructions for the user."""
    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║          ✅ DASHBOARD PROJECT READY TO OPEN                 ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()
    print(f"  Project location: {project_root}")
    print(f"  Project file:     {project_root / f'{project_name}.pbip'}")
    print()
    print("  ┌─────────────────────────────────────────────────────────┐")
    print("  │  NEXT STEPS:                                            │")
    print("  │  1. Open Power BI Desktop                               │")
    print("  │  2. File → Open → select the .pbip file                 │")
    print("  │  3. Wait for all visuals to load                        │")
    print("  │  4. Check each page for correct rendering               │")
    print("  │  5. Adjust visual data bindings if needed               │")
    print("  │  6. Publish to Power BI Service if desired              │")
    print("  └─────────────────────────────────────────────────────────┘")
    print()
    print("  ⚠️  IMPORTANT REMINDERS:")
    print("  • If visuals don't load: close PBID, delete cache.abf, reopen")
    print("  • If measures don't appear: PBID recalculates on open")
    print(f"  • Theme: {BUILTIN_THEMES.get('4', {}).get('name', 'applied')}")
    print()
    print("  📋 Project structure:")
    for item in sorted(project_root.rglob("*")):
        if item.is_file() and item.name != "cache.abf":
            rel = item.relative_to(project_root)
            print(f"    {rel}")
    print()


# ──────────────────────────────────────────────────────────────
# MAIN ORCHESTRATOR
# ──────────────────────────────────────────────────────────────

def main():
    banner()

    args = parse_args()

    # Determine project path
    project_root = None
    project_name = None
    data_file = None

    if args.create:
        # Create mode: user provides name and data file
        project_name = args.create[0]
        data_file = Path(args.create[1])
        if not data_file.exists():
            error(f"Data file not found: {data_file}")
            sys.exit(1)

        # Ask user to create the PBIP project first
        project_root = Path.cwd() / project_name
        info(f"Project will be created at: {project_root}")
        info("Before continuing:")
        info(f"  1. Open Power BI Desktop (Blank Report)")
        info(f"  2. Get Data → CSV/Excel → select '{data_file.name}'")
        info(f"  3. Load the data (or Close & Apply)")
        info(f"  4. File → Save As → Power BI Project (*.pbip) → name it '{project_name}'")
        info(f"  5. Save to: {project_root}")
        info(f"  6. Close Power BI Desktop")
        if not confirm("Have you completed the above steps?", default="no"):
            warn("Please complete the data loading in PBID, save as PBIP, and re-run.")
            info("Command for next run: python new_powerbi_dashboard.py <project.pbip>")
            sys.exit(1)

        # Verify structure
        if not verify_project_structure(project_root, project_name):
            warn("Project structure incomplete. Please ensure the .pbip was saved correctly.")
            if not confirm("Continue anyway?", default="no"):
                sys.exit(1)
    elif args.project:
        # Use provided project path
        pbip_path = Path(args.project)
        if not pbip_path.exists():
            # Maybe it's a directory name without .pbip
            candidates = list(Path(args.project).rglob("*.pbip"))
            if candidates:
                pbip_path = candidates[0]
            elif Path(f"{args.project}.pbip").exists():
                pbip_path = Path(f"{args.project}.pbip")
            elif Path(args.project).is_dir():
                candidates = list(Path(args.project).glob("*.pbip"))
                if candidates:
                    pbip_path = candidates[0]
                else:
                    error(f"No .pbip file found at: {args.project}")
                    sys.exit(1)
            else:
                # Try relative
                rel = Path.cwd() / f"{args.project}.pbip"
                if rel.exists():
                    pbip_path = rel
                else:
                    error(f"Project not found: {args.project}")
                    sys.exit(1)

        project_root = pbip_path.parent
        project_name = pbip_path.stem
        info(f"Project: {project_name} at {project_root}")
    else:
        # Search for a .pbip in current directory
        candidates = list(Path.cwd().glob("*.pbip"))
        if not candidates:
            error("No .pbip file found. Provide a project path or use --create.")
            error("Usage: python new_powerbi_dashboard.py <project.pbip>")
            error("   or: python new_powerbi_dashboard.py --create MyDashboard data.csv")
            sys.exit(1)
        pbip_path = candidates[0]
        project_root = pbip_path.parent
        project_name = pbip_path.stem
        info(f"Found project: {project_name} at {project_root}")

    # =================================================================
    # PHASE 1: PREFLIGHT
    # =================================================================
    step("PHASE 1: PROJECT PREPARATION")

    if not check_pbir_cli():
        error("pbir CLI is required. Install with: pip install pbir-cli")
        sys.exit(1)

    verify_project_structure(project_root, project_name)
    close_pbid()
    delete_cache(project_root, project_name)
    run_validation(project_root, project_name)

    # =================================================================
    # PHASE 2 & 3: DATA ANALYSIS + USER PREFERENCES
    # =================================================================
    analysis = None
    if not args.skip_data_analysis and data_file:
        analysis = analyze_data_file(data_file)
    elif not args.skip_data_analysis and not args.non_interactive:
        data_file_path = ask("What is the path to your data file (.csv or .xlsx)?")
        if data_file_path:
            analysis = analyze_data_file(data_file_path)
            if analysis:
                dims, measures, dates = suggest_dimensions_measures(analysis, "analysis")
                if dims:
                    info(f"\n  Suggested dimensions: {', '.join(dims[:6])}")
                if measures:
                    info(f"  Suggested measures: {', '.join(measures[:6])}")
                if dates:
                    info(f"  Date columns: {', '.join(dates[:3])}")
    elif not args.skip_data_analysis:
        info("Non-interactive mode: attempting to find data file in project...")
        # Try to find TMDL columns instead
        table_name, columns = read_table_tmdl(project_root, project_name)
        if table_name:
            info(f"Using table '{table_name}' with columns from TMDL")

    # Collect user preferences
    prefs = collect_user_preferences(args, analysis)

    # =================================================================
    # PHASE 4: DAX MEASURES
    # =================================================================
    generate_and_add_measures(
        project_root, project_name,
        analysis, prefs["purpose"],
        args.non_interactive
    )

    # =================================================================
    # PHASE 5: VISUALS
    # =================================================================
    create_visuals(
        project_root, project_name,
        prefs["num_pages"],
        analysis, prefs["purpose"],
        args.non_interactive
    )

    # =================================================================
    # PHASE 6: THEME
    # =================================================================
    # Ensure PBID closed before theme
    close_pbid()
    delete_cache(project_root, project_name)

    apply_theme(project_root, project_name, prefs["theme_choice"])

    # =================================================================
    # PHASE 7: FINAL VERIFICATION
    # =================================================================
    final_verification(project_root, project_name)

    # Done!
    print_final_instructions(project_root, project_name)

    info("All phases complete!")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n\nInterrupted by user. Exiting.")
        sys.exit(130)
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
